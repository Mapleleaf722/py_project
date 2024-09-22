import vobject
from openpyxl import Workbook
import re


def read_vcf(file_path):
    with open(file_path, 'r') as file:
        vcard_string = file.read()
        pattern = r"BEGIN:VCARD"
        result = re.split(pattern, vcard_string)
        vcard = []
        for i in result[1:]:
            i = "BEGIN:VCARD" + i
            try:
                vcard.append(vobject.readOne(i))
                print(i)
            except:
                pass
    print(vcard)
    return vcard


def write_excel(vcard, output_file):
    workbook = Workbook()
    sheet = workbook.active

    # 写入表头
    sheet.cell(row=1, column=1, value='姓名')
    sheet.cell(row=1, column=2, value='电话号码')
    print(str(vcard))
    # 写入联系人信息
    row = 2
    for i in vcard:
        try:
            for contact in i.contents['fn']:
                sheet.cell(row=row, column=1, value=str(contact.value))

            for contact in i.contents['tel']:
                sheet.cell(row=row, column=2, value=str(contact.value))
        except:
            pass
        row += 1

    workbook.save(output_file)


if __name__ == '__main__':
    # 读取vcf文件
    vcf = read_vcf(r'D:\Files\info\Contacts-2024-07-09.vcf')
    # 转换为Excel文件
    write_excel(vcf, 'contacts.xlsx')
